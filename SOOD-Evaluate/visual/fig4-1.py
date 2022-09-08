from logging import handlers
import torch
import matplotlib.pyplot as plt
from pylab import *
from openpyxl import load_workbook
import matplotlib as mpl

def get_data(start_row):
    x = [1,5,10,15,20,25]
    y1,y2,y3,y4=[],[],[],[]
    for i in range(col_num):
        y1.append(ws[start_row][i+1].value)
        y2.append(ws[start_row+1][i+1].value)
        y3.append(ws[start_row+2][i+1].value)
        y4.append(ws[start_row+3][i+1].value)
    return x,y1,y2,y3,y4

mpl.use('Agg')

#plot set
font_size=6
marker_size='2'
fig_size=(7.5,1.5)
fig_dpi=300
set_alpha=1
colors=['#F9A968','#4D85BD']
fig_name='fig4-1'
col_num=6

#data source
show_data=dict()
wb = load_workbook('./visual/visual.xlsx')
ws = wb[fig_name]

#plot fig
fig=plt.figure(figsize=fig_size,dpi=fig_dpi)
subplots_adjust(wspace=0.3,bottom=0.3,left=0.05)

x,y1,y2,y3,y4=get_data(2)
ax1=plt.subplot(151)
plt.ylim(68,83)
plt.xticks(fontsize=font_size,ticks=x)
plt.yticks([70,75,80],fontsize=font_size)
ax1.tick_params(which='major',width=0.3,length=1,pad=1)
ax1.set_xlabel(xlabel='Number of classifiers',fontsize=font_size,labelpad=1)
ax1.set_ylabel(ylabel='Accuracy',fontsize=font_size,labelpad=1)
plt.title('(a) ID, Accuracy',fontsize=font_size)
ax1 = plt.gca()
ax1.spines['top'].set_linewidth(0)
ax1.spines['bottom'].set_linewidth(0.2)
ax1.spines['left'].set_linewidth(0.2)
ax1.spines['right'].set_linewidth(0)

ax1.plot(x,y1,color=colors[0],linestyle='--',alpha=set_alpha,linewidth=1,marker='^',markeredgecolor=colors[0],markersize=marker_size,markeredgewidth=1,\
    label='PlainEN')
ax1.plot(x,y2,color=colors[0],linestyle='-',alpha=set_alpha, linewidth=1,marker='^',markeredgecolor=colors[0],markersize=marker_size,markeredgewidth=1,\
    label='LSREN')
ax1.plot(x,y3,color=colors[1],linestyle='--',alpha=set_alpha, linewidth=1,marker='o',markeredgecolor=colors[1],markersize=marker_size,markeredgewidth=1,\
    label='PlainEN+')
ax1.plot(x,y4,color=colors[1],linestyle='-',alpha=set_alpha, linewidth=1,marker='o',markeredgecolor=colors[1],markersize=marker_size,markeredgewidth=1,\
    label='LSREN+')
# fig.legend(fontsize=font_size,frameon=False,bbox_to_anchor=(1.1,1.0))


x,y1,y2,y3,y4=get_data(8)
ax2=plt.subplot(152)
plt.ylim(68,83)
plt.xticks(fontsize=font_size,ticks=x)
plt.yticks([70,75,80],fontsize=font_size)
ax2.tick_params(which='major',width=0.3,length=1,pad=1)
ax2.set_xlabel(xlabel='Number of classifiers',fontsize=font_size,labelpad=1)
ax2.set_ylabel(ylabel='AUROC',fontsize=font_size,labelpad=1)
plt.title('(b) SOOD, AUROC',fontsize=font_size)
ax2 = plt.gca()
ax2.spines['top'].set_linewidth(0)
ax2.spines['bottom'].set_linewidth(0.2)
ax2.spines['left'].set_linewidth(0.2)
ax2.spines['right'].set_linewidth(0)

ax2.plot(x,y1,color=colors[0],linestyle='--',alpha=set_alpha,linewidth=1,marker='^',markeredgecolor=colors[0],markersize=marker_size,markeredgewidth=1,\
    label='PlainEN')
ax2.plot(x,y2,color=colors[0],linestyle='-',alpha=set_alpha, linewidth=1,marker='^',markeredgecolor=colors[0],markersize=marker_size,markeredgewidth=1,\
    label='LSREN')
ax2.plot(x,y3,color=colors[1],linestyle='--',alpha=set_alpha, linewidth=1,marker='o',markeredgecolor=colors[1],markersize=marker_size,markeredgewidth=1,\
    label='PlainEN+')
ax2.plot(x,y4,color=colors[1],linestyle='-',alpha=set_alpha, linewidth=1,marker='o',markeredgecolor=colors[1],markersize=marker_size,markeredgewidth=1,\
    label='LSREN+')

x,y1,y2,y3,y4=get_data(14)
ax3=plt.subplot(153)

plt.ylim(8,35)
plt.xticks(fontsize=font_size,ticks=x)
plt.yticks([10,15,20,25,30],fontsize=font_size)
ax3.tick_params(which='major',width=0.3,length=1,pad=1)
ax3.set_xlabel(xlabel='Number of classifiers',fontsize=font_size,labelpad=1)
ax3.set_ylabel(ylabel='TNR@TPR95',fontsize=font_size,labelpad=1)
plt.title('(c) SOOD, TNR@TPR95',fontsize=font_size)
ax3 = plt.gca()
ax3.spines['top'].set_linewidth(0)
ax3.spines['bottom'].set_linewidth(0.2)
ax3.spines['left'].set_linewidth(0.2)
ax3.spines['right'].set_linewidth(0)

ax3.plot(x,y1,color=colors[0],linestyle='--',alpha=set_alpha,linewidth=1,marker='^',markeredgecolor=colors[0],markersize=marker_size,markeredgewidth=1,\
    label='PlainEN')
ax3.plot(x,y2,color=colors[0],linestyle='-',alpha=set_alpha, linewidth=1,marker='^',markeredgecolor=colors[0],markersize=marker_size,markeredgewidth=1,\
    label='LSREN')
ax3.plot(x,y3,color=colors[1],linestyle='--',alpha=set_alpha, linewidth=1,marker='o',markeredgecolor=colors[1],markersize=marker_size,markeredgewidth=1,\
    label='PlainEN+')
ax3.plot(x,y4,color=colors[1],linestyle='-',alpha=set_alpha, linewidth=1,marker='o',markeredgecolor=colors[1],markersize=marker_size,markeredgewidth=1,\
    label='LSREN+')


x,y1,y2,y3,y4=get_data(20)
ax4=plt.subplot(154)
plt.ylim(75,95)
plt.xticks(fontsize=font_size,ticks=x)
plt.yticks([70,75,80,85,90],fontsize=font_size)
ax4.tick_params(which='major',width=0.3,length=1,pad=1)
ax4.set_xlabel(xlabel='Number of classifiers',fontsize=font_size,labelpad=1)
ax4.set_ylabel(ylabel='AUROC',fontsize=font_size,labelpad=1)
plt.title('(d) OOD, AUROC',fontsize=font_size)
ax4 = plt.gca()
ax4.spines['top'].set_linewidth(0)
ax4.spines['bottom'].set_linewidth(0.2)
ax4.spines['left'].set_linewidth(0.2)
ax4.spines['right'].set_linewidth(0)

ax4.plot(x,y1,color=colors[0],linestyle='--',alpha=set_alpha,linewidth=1,marker='^',markeredgecolor=colors[0],markersize=marker_size,markeredgewidth=1,\
    label='PlainEN')
ax4.plot(x,y2,color=colors[0],linestyle='-',alpha=set_alpha, linewidth=1,marker='^',markeredgecolor=colors[0],markersize=marker_size,markeredgewidth=1,\
    label='LSREN')
ax4.plot(x,y3,color=colors[1],linestyle='--',alpha=set_alpha, linewidth=1,marker='o',markeredgecolor=colors[1],markersize=marker_size,markeredgewidth=1,\
    label='PlainEN+')
ax4.plot(x,y4,color=colors[1],linestyle='-',alpha=set_alpha, linewidth=1,marker='o',markeredgecolor=colors[1],markersize=marker_size,markeredgewidth=1,\
    label='LSREN+')

x,y1,y2,y3,y4=get_data(26)
ax5=plt.subplot(155)
plt.ylim(15,70)
plt.xticks(fontsize=font_size,ticks=x)
plt.yticks([20,30,40,50,60],fontsize=font_size)
ax5.tick_params(which='major',width=0.3,length=1,pad=1)
ax5.set_xlabel(xlabel='Number of classifiers',fontsize=font_size,labelpad=1)
ax5.set_ylabel(ylabel='TNR@TPR95',fontsize=font_size,labelpad=1)
plt.title('(e)OOD, TNR@TPR95',fontsize=font_size)
ax5 = plt.gca()
ax5.spines['top'].set_linewidth(0)
ax5.spines['bottom'].set_linewidth(0.2)
ax5.spines['left'].set_linewidth(0.2)
ax5.spines['right'].set_linewidth(0)

line1,=ax5.plot(x,y1,color=colors[0],linestyle='--',alpha=set_alpha,linewidth=1,marker='^',markeredgecolor=colors[0],markersize=marker_size,markeredgewidth=1,\
    label='PlainEN')
line2,=ax5.plot(x,y2,color=colors[0],linestyle='-',alpha=set_alpha, linewidth=1,marker='^',markeredgecolor=colors[0],markersize=marker_size,markeredgewidth=1,\
    label='LSREN')
line3,=ax5.plot(x,y3,color=colors[1],linestyle='--',alpha=set_alpha, linewidth=1,marker='o',markeredgecolor=colors[1],markersize=marker_size,markeredgewidth=1,\
    label='PlainEN+')
line4,=ax5.plot(x,y4,color=colors[1],linestyle='-',alpha=set_alpha, linewidth=1,marker='o',markeredgecolor=colors[1],markersize=marker_size,markeredgewidth=1,\
    label='LSREN+')

fig.legend(handles=[line1,line2,line3,line4], fontsize=font_size,frameon=False,handlelength=0.8,bbox_to_anchor=(1.0,0.98))


plt.savefig('./visual/fig/{}.png'.format(fig_name))
plt.savefig('./visual/fig/{}.eps'.format(fig_name),format='eps')
plt.close()


