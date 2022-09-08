import matplotlib.pyplot as plt
import matplotlib as mpl
from openpyxl import load_workbook
from pylab import *
import matplotlib.patches as mpatches
mpl.use('Agg')

def get_visual_data(row,col_num):
    x=list(range(col_num))
    y1,y2=[],[]
    for i in range(col_num):
        y1.append(ws[row][i+1].value)
        y2.append(ws[row+1][i+1].value)
    return x,y1,y2

interval_x=0.4
width_x=0.4
text_space=0.5
font_size=6
set_alpha=1
set_dpi=300
set_figsize=[7.2,4]
colors=['#F9A968','#4D85BD']

fig_name='fig4-2'
start_row,col_num=2,7
# colors = ["#51C1C8", "#E96279", "#44A2D6", "#536D84","#51C1C8", "#E96279", "#44A2D6", "#536D84"]


show_data=dict()
wb = load_workbook('./visual/visual.xlsx')
ws = wb[fig_name]



fig=plt.figure(dpi=set_dpi,figsize=set_figsize)
subplots_adjust(left=0.05,hspace=0.4)

ax=plt.subplot(313)
x,y1,y2=get_visual_data(start_row,col_num)
plt.title('(c) ID data, Accuracy',fontsize=font_size)
tick_params(which='major',width=0.3,length=1)
plt.xticks([])
plt.ylim(68,82)
plt.yticks(fontsize=font_size,ticks=[70,75,80])
plt.ylabel('Accuracy',fontsize=font_size)
plt.bar(x, y1, color=colors[0],width=width_x,align='center',label='w/o OODBN',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y1[i]+text_space,'{:.1f}'.format(y1[i]),fontsize=font_size,ha='center')
for i in range(len(x)):x[i]=x[i]+interval_x
plt.bar(x, y2, color=colors[1],width=width_x,align='center',label='with OODBN',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y2[i]+text_space,'{:.1f}'.format(y2[i]),fontsize=font_size,ha='center')
[ax.spines[item].set_linewidth(0) for item in ['top','right']]
[ax.spines[item].set_linewidth(0.5) for item in ['left','bottom']]
for i in range(len(x)):x[i]=x[i]-0.5*interval_x
plt.xticks(x,['Baseline','ODIN','GODIN','Dropout-34','Dropout-50','PlainEN','LSREN'],fontsize=font_size)

ax=plt.subplot(311)
x,y1,y2=get_visual_data(start_row+4,col_num)
plt.title('(a) SOOD data, AUROC',fontsize=font_size)
tick_params(which='major',width=0.3,length=1)
plt.xticks([])
plt.ylim(55,85)
plt.yticks(fontsize=font_size,ticks=[60,70,80])
plt.ylabel('AUROC',fontsize=font_size)
plt.bar(x, y1, color=colors[0],width=width_x,align='center',label='w/o OODBN',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y1[i]+text_space,'{:.1f}'.format(y1[i]),fontsize=font_size,ha='center')
for i in range(len(x)):x[i]=x[i]+interval_x
plt.bar(x, y2, color=colors[1],width=width_x,align='center',label='with OODBN',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y2[i]+text_space,'{:.1f}'.format(y2[i]),fontsize=font_size,ha='center')
[ax.spines[item].set_linewidth(0) for item in ['top','right']]
[ax.spines[item].set_linewidth(0.5) for item in ['left','bottom']]

# ax=plt.subplot(513)
# x,y1,y2=get_visual_data(start_row+8,col_num)
# plt.title('(c) SOOD data, TNR@TPR95',fontsize=font_size)
# tick_params(which='major',width=0.3,length=1)
# plt.xticks([])
# plt.ylim(5,35)
# plt.yticks(fontsize=font_size,ticks=[10,20,30])
# plt.ylabel('TNR@TPR95',fontsize=font_size)
# plt.bar(x, y1, color=colors[0],width=width_x,align='center',label='w/o osbn',alpha=set_alpha)
# for i in range(len(x)):
#     plt.text(x[i],y1[i]+text_space,'{:.1f}'.format(y1[i]),fontsize=font_size,ha='center')
# for i in range(len(x)):x[i]=x[i]+interval_x
# plt.bar(x, y2, color=colors[1],width=width_x,align='center',label='with osbn',alpha=set_alpha)
# for i in range(len(x)):
#     plt.text(x[i],y2[i]+text_space,'{:.1f}'.format(y2[i]),fontsize=font_size,ha='center')
# [ax.spines[item].set_linewidth(0) for item in ['top','right']]
# [ax.spines[item].set_linewidth(0.5) for item in ['left','bottom']]

ax=plt.subplot(312)
x,y1,y2=get_visual_data(start_row+12,col_num)
plt.title('(b) OOD data, AUROC',fontsize=font_size)
tick_params(which='major',width=0.3,length=1)
plt.xticks([])
plt.ylim(65,95)
plt.yticks(fontsize=font_size,ticks=[70,80,90])
plt.ylabel('AUROC',fontsize=font_size)
plt.bar(x, y1, color=colors[0],width=width_x,align='center',label='w/o OODBN',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y1[i]+text_space,'{:.1f}'.format(y1[i]),fontsize=font_size,ha='center')
for i in range(len(x)):x[i]=x[i]+interval_x
plt.bar(x, y2, color=colors[1],width=width_x,align='center',label='with OODBN',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y2[i]+text_space,'{:.1f}'.format(y2[i]),fontsize=font_size,ha='center')
[ax.spines[item].set_linewidth(0) for item in ['top','right']]
[ax.spines[item].set_linewidth(0.5) for item in ['left','bottom']]

# ax=plt.subplot(515)
# x,y1,y2=get_visual_data(start_row+16,col_num)
# plt.title('(e) SOOD data, TNR@TPR95',fontsize=font_size)
# tick_params(which='major',width=0.3,length=1)
# plt.xticks([])
# plt.ylim(10,60)
# plt.yticks(fontsize=font_size,ticks=[15,35,55])
# plt.ylabel('TNR@TPR95',fontsize=font_size)
# plt.bar(x, y1, color=colors[0],width=width_x,align='center',label='w/o osbn',alpha=set_alpha)
# for i in range(len(x)):
#     plt.text(x[i],y1[i]+text_space,'{:.1f}'.format(y1[i]),fontsize=font_size,ha='center')
# for i in range(len(x)):x[i]=x[i]+interval_x
# plt.bar(x, y2, color=colors[1],width=width_x,align='center',label='with osbn',alpha=set_alpha)
# for i in range(len(x)):
#     plt.text(x[i],y2[i]+text_space,'{:.1f}'.format(y2[i]),fontsize=font_size,ha='center')
# [ax.spines[item].set_linewidth(0) for item in ['top','right']]
# [ax.spines[item].set_linewidth(0.5) for item in ['left','bottom']]




recs=[]
for i in range(2):
    recs.append(mpatches.Rectangle((0,0),0.2,0.2,fc=colors[i]))
fig.legend(recs,['w/o OODBN','with OODBN'],fontsize=font_size,frameon=False,handletextpad=0.2,\
    handlelength=0.8,bbox_to_anchor=(0.98,0.90))
     
plt.savefig('./visual/fig/{}.png'.format(fig_name))
plt.savefig('./visual/fig/{}.eps'.format(fig_name),format='eps')
# plt.show()
plt.close()