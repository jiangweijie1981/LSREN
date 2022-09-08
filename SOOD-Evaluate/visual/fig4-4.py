import matplotlib.pyplot as plt
import matplotlib as mpl
from openpyxl import load_workbook
from pylab import *
import matplotlib.patches as mpatches
mpl.use('Agg')

def get_visual_data(row,col_num):
    x=list(range(col_num))
    y1,y2,y3=[],[],[]
    for i in range(col_num):
        y1.append(ws[row][i+1].value)
        y2.append(ws[row+1][i+1].value)
        y3.append(ws[row+2][i+1].value)
    return x,y1,y2,y3

interval_x=0.25
width_x=0.20
text_space=0.5
font_size=6
set_alpha=1
set_dpi=300
set_figsize=[7.5,2]
colors=['#4D85BD','#85C085','#F9A968']

fig_name='fig4-4'
start_row,col_num=2,2
# colors = ["#51C1C8", "#E96279", "#44A2D6", "#536D84","#51C1C8", "#E96279", "#44A2D6", "#536D84"]


show_data=dict()
wb = load_workbook('./visual/visual.xlsx')
ws = wb[fig_name]



fig=plt.figure(dpi=set_dpi,figsize=set_figsize)
subplots_adjust(hspace=1,left=0.05)

ax=plt.subplot(221)
x,y1,y2,y3=get_visual_data(start_row,col_num)
plt.title('(a) LSREN, SOOD data',fontsize=font_size)
tick_params(which='major',width=0.3,length=1)
plt.xticks([])
plt.ylim(40,95)
plt.yticks(fontsize=font_size,ticks=[50,70,90])
plt.ylabel('AUROC',fontsize=font_size)
plt.bar(x, y1, color=colors[0],width=width_x,align='center',label='Mean',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y1[i]+text_space,'{:.1f}'.format(y1[i]),fontsize=font_size,ha='center')
for i in range(len(x)):x[i]=x[i]+interval_x
plt.bar(x, y2, color=colors[1],width=width_x,align='center',label='Deviation',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y2[i]+text_space,'{:.1f}'.format(y2[i]),fontsize=font_size,ha='center')
for i in range(len(x)):x[i]=x[i]+interval_x
plt.bar(x, y3, color=colors[2],width=width_x,align='center',label='Both',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y3[i]+text_space,'{:.1f}'.format(y3[i]),fontsize=font_size,ha='center')
[ax.spines[item].set_linewidth(0) for item in ['top','right']]
[ax.spines[item].set_linewidth(0.5) for item in ['left','bottom']]
for i in range(len(x)):x[i]=x[i]-interval_x
plt.xticks(x,['Without logits smoothing','With logits smoothing'],fontsize=font_size)

ax=plt.subplot(222)
x,y1,y2,y3=get_visual_data(start_row+10,col_num)
plt.title('(b) LSREN+, SOOD data',fontsize=font_size)
tick_params(which='major',width=0.3,length=1)
plt.xticks([])
plt.ylim(40,95)
plt.yticks(fontsize=font_size,ticks=[50,70,90])
plt.ylabel('AUROC',fontsize=font_size)
plt.bar(x, y1, color=colors[0],width=width_x,align='center',label='Mean',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y1[i]+text_space,'{:.1f}'.format(y1[i]),fontsize=font_size,ha='center')
for i in range(len(x)):x[i]=x[i]+interval_x
plt.bar(x, y2, color=colors[1],width=width_x,align='center',label='Deviation',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y2[i]+text_space,'{:.1f}'.format(y2[i]),fontsize=font_size,ha='center')
for i in range(len(x)):x[i]=x[i]+interval_x
plt.bar(x, y3, color=colors[2],width=width_x,align='center',label='Both',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y3[i]+text_space,'{:.1f}'.format(y3[i]),fontsize=font_size,ha='center')
[ax.spines[item].set_linewidth(0) for item in ['top','right']]
[ax.spines[item].set_linewidth(0.5) for item in ['left','bottom']]

for i in range(len(x)):x[i]=x[i]-interval_x
plt.xticks(x,['Without logits smoothing','With logits smoothing'],fontsize=font_size)



ax=plt.subplot(223)
x,y1,y2,y3=get_visual_data(start_row+5,col_num)
plt.title('(c) LSREN, OOD data',fontsize=font_size)
tick_params(which='major',width=0.3,length=1)
plt.xticks([])
plt.ylim(40,95)
plt.yticks(fontsize=font_size,ticks=[50,70,90])
plt.ylabel('AUROC',fontsize=font_size)
plt.bar(x, y1, color=colors[0],width=width_x,align='center',label='Mean',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y1[i]+text_space,'{:.1f}'.format(y1[i]),fontsize=font_size,ha='center')
for i in range(len(x)):x[i]=x[i]+interval_x
plt.bar(x, y2, color=colors[1],width=width_x,align='center',label='Deviation',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y2[i]+text_space,'{:.1f}'.format(y2[i]),fontsize=font_size,ha='center')
for i in range(len(x)):x[i]=x[i]+interval_x
plt.bar(x, y3, color=colors[2],width=width_x,align='center',label='Both',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y3[i]+text_space,'{:.1f}'.format(y3[i]),fontsize=font_size,ha='center')
[ax.spines[item].set_linewidth(0) for item in ['top','right']]
[ax.spines[item].set_linewidth(0.5) for item in ['left','bottom']]
for i in range(len(x)):x[i]=x[i]-interval_x
plt.xticks(x,['Without logits smoothing','With logits smoothing'],fontsize=font_size)

ax=plt.subplot(224)
x,y1,y2,y3=get_visual_data(start_row+15,col_num)
plt.title('(d) LSREN+, OOD data',fontsize=font_size)
tick_params(which='major',width=0.3,length=1)
plt.xticks([])
plt.ylim(40,95)
plt.yticks(fontsize=font_size,ticks=[50,70,90])
plt.ylabel('AUROC',fontsize=font_size)
plt.bar(x, y1, color=colors[0],width=width_x,align='center',label='Mean',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y1[i]+text_space,'{:.1f}'.format(y1[i]),fontsize=font_size,ha='center')
for i in range(len(x)):x[i]=x[i]+interval_x
plt.bar(x, y2, color=colors[1],width=width_x,align='center',label='Deviation',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y2[i]+text_space,'{:.1f}'.format(y2[i]),fontsize=font_size,ha='center')
for i in range(len(x)):x[i]=x[i]+interval_x
plt.bar(x, y3, color=colors[2],width=width_x,align='center',label='Both',alpha=set_alpha)
for i in range(len(x)):
    plt.text(x[i],y3[i]+text_space,'{:.1f}'.format(y3[i]),fontsize=font_size,ha='center')
[ax.spines[item].set_linewidth(0) for item in ['top','right']]
[ax.spines[item].set_linewidth(0.5) for item in ['left','bottom']]

for i in range(len(x)):x[i]=x[i]-interval_x
plt.xticks(x,['Without logits smoothing','With logits smoothing'],fontsize=font_size)


recs=[]
for i in range(3):
    recs.append(mpatches.Rectangle((0,0),0.2,0.2,fc=colors[i]))
fig.legend(recs,['Mean','Deviation','Both'],fontsize=font_size,frameon=False,handletextpad=0.2,\
    handlelength=0.8,bbox_to_anchor=(0.98,0.90))
     
plt.savefig('./visual/fig/{}.png'.format(fig_name))
plt.savefig('./visual/fig/{}.eps'.format(fig_name),format='eps')
# plt.show()
plt.close()