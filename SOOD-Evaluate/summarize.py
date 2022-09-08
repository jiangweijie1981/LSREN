from torch import seed
from configs import *
from utils import *
import os
from cal_metric import *
import shutil
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


if __name__=='__main__':
    args=Config().data
    table_list=['Pred','Pred_Tsc']
    ood,sood=dict(),dict()
    # ood['mean'],ood['bias'],ood['cali']=[],[],[]
    # sood['mean'],sood['bias'],sood['cali']=[],[],[]

    # for s in [0,3,4]:
    #     args.seed=s
    #     set_seed(args.seed)
    #     ensemble_path = '../models/{}_seed{}-{}_{}-{}-{}/{}_esbl{}_range{}/ensemble_epochs{}'.format(
    #         args.net,args.seed,args.seed,args.in_dataset,args.dataset_id,args.samples_pre_class_num,
    #         args.train_type,args.ensemble_num,args.smooth_range,args.epochs)

    #     wb = load_workbook('{}/{}_{}_{}.xlsx'.format(ensemble_path,args.inferance[0],args.in_dataset,args.ood_batch_size))
    #     ws = wb['Pred']
    #     ood_aur.append((ws[7][1].value+ws[9][1].value+ws[11][1].value+ws[13][1].value+ws[15][1].value)/5)
    #     ood_tnr.append((ws[6][1].value+ws[8][1].value+ws[10][1].value+ws[12][1].value+ws[14][1].value)/5)
    #     sood_aur.append(ws[3][1].value)
    #     sood_tnr.append(ws[2][1].value)
    

    ensemble_path = '../models/{}_seed{}-{}_{}-{}-{}/{}_esbl{}_range{}/ensemble_epochs{}'.format(
        args.net,args.seed_begin,args.seed_end,args.in_dataset,args.dataset_id,args.samples_pre_class_num,
        args.train_type,args.ensemble_num,args.smooth_range,args.epochs)

    # ensemble_path = '../models/{}_seed{}_{}_{}-{}/{}_esbl{}_range{}/ensemble_epochs{}'.format(
    #     args.net,args.seed,args.in_dataset,args.dataset_id,args.samples_pre_class_num,
    #     args.train_type,args.ensemble_num,args.smooth_range,args.epochs)

    wb = load_workbook('{}/{}_{}_{}.xlsx'.format(ensemble_path,args.inferance[0],args.in_dataset,args.ood_batch_size))
    # ws = wb['Pred_Tsc']
    ws = wb['Pred']

    max_col,max_value=1,ws[4][1].value+ws[5][1].value
    for i in range(21):
        cur_value=ws[4][i+1].value+ws[5][i+1].value
        if cur_value>max_value:
            max_col=i+1
            max_value=cur_value

    start_row=2
    sood['mean']=(ws[start_row][1].value)
    sood['bias']=(ws[start_row][21].value)
    sood['cali']=(ws[start_row][max_col].value)

    start_row+=4
    ood['mean']=((ws[start_row][1].value+ws[start_row+2][1].value+ws[start_row+4][1].value+ws[start_row+6][1].value+ws[start_row+8][1].value)/5)
    ood['bias']=((ws[start_row][21].value+ws[start_row+2][21].value+ws[start_row+4][21].value+ws[start_row+6][21].value+ws[start_row+8][21].value)/5)
    ood['cali']=((ws[start_row][max_col].value+ws[start_row+2][max_col].value+ws[start_row+4][max_col].value+ws[start_row+6][max_col].value+ws[start_row+8][max_col].value)/5)

    with open('summarize_en{}_{}_{}_{}.txt'.format(args.ensemble_num,args.ood_batch_size,args.seed_end,start_row),'wt') as f:
        print('ood:',file=f)
        print('mean:{:.1f}, bias:{:.1f}, cali:{:.1f}\n'.format(ood['mean'],ood['bias'],ood['cali']),file=f)
        print('sood',file=f)
        print('mean:{:.1f}, bias:{:.1f}, cali:{:.1f}\n'.format(sood['mean'],sood['bias'],sood['cali']),file=f)





