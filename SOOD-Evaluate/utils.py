import numpy as np
import random
import torch
import os

def set_seed(seed):
    np.random.seed(seed)
    random.seed(seed)
    torch.manual_seed(seed) # cpu
    torch.cuda.manual_seed_all(seed)  # gpu
    torch.backends.cudnn.benchmark=False
    torch.backends.cudnn.deterministic = True

def get_smooth_list(args):
    smooth_list=[]
    if args.ensemble_num==1:
        smooth_list.append(0.0)
    elif args.ensemble_num==2:
        smooth_list.append(args.smooth_range)
    else:
        for i in range(args.ensemble_num-2):
            smooth_list.append(round(args.smooth_range*(i+1)/(args.ensemble_num-1),6))
    return smooth_list

def get_ensemble_list(args):
    model_file_list=[]
    smooth_list=get_smooth_list(args)
    if args.train_type=='LSR':
        for i in range(len(smooth_list)):
            args.smooth=smooth_list[i]
            file_path = '../models/{}_seed{}_{}_{}-{}/{}_esbl{}_range{}/epochs{}_s{:.6f}'.format(
                args.net,args.seed,args.in_dataset,args.dataset_id,args.samples_pre_class_num,
                args.train_type,args.ensemble_num,args.smooth_range,args.epochs,args.smooth)
            model_file_list.append(file_path)
        if args.ensemble_num>1:
            args.smooth=0.0
            file_path = '../models/{}_seed{}_{}_{}-{}/{}_esbl1_range{}/epochs{}_s{:.6f}'.format(
                args.net,args.seed,args.in_dataset,args.dataset_id,args.samples_pre_class_num,
                args.train_type,args.smooth_range,args.epochs,args.smooth)
            model_file_list.append(file_path)
        if args.ensemble_num>2:
            args.smooth=args.smooth_range
            file_path = '../models/{}_seed{}_{}_{}-{}/{}_esbl2_range{}/epochs{}_s{:.6f}'.format(
                args.net,args.seed,args.in_dataset,args.dataset_id,args.samples_pre_class_num,
                args.train_type,args.smooth_range,args.epochs,args.smooth)
            model_file_list.append(file_path)
    return model_file_list
def get_net(args):
    if args.net=='resnet':
        import net.resnet as resnet
        net = resnet.ResNet34(num_c=args.class_num)
    elif args.net=='resnet_cos':
        import net.resnet_cos as resnet
        net = resnet.ResNet34(num_c=args.class_num)
    elif args.net=='resnet_inn':
        import net.resnet_inn as resnet
        net = resnet.ResNet34(num_c=args.class_num)
    elif args.net=='resnet_euc':
        import net.resnet_euc as resnet
        net = resnet.ResNet34(num_c=args.class_num)
    elif args.net=='resnet_dropout34':
        import net.resnet_dropout as resnet
        net = resnet.ResNet34(num_c=args.class_num)
    elif args.net=='resnet_dropout50':
        import net.resnet_dropout as resnet
        net = resnet.ResNet50(num_c=args.class_num)
    if torch.cuda.device_count()>1:
        net=torch.nn.DataParallel(net)
    return net.cuda()
def load_net(args,model_path):
    if args.net=='resnet':
        import net.resnet as resnet
        net = resnet.ResNet34(num_c=args.class_num)
    elif args.net=='resnet_cos':
        import net.resnet_cos as resnet
        net = resnet.ResNet34(num_c=args.class_num)
    elif args.net=='resnet_inn':
        import net.resnet_inn as resnet
        net = resnet.ResNet34(num_c=args.class_num)
    elif args.net=='resnet_euc':
        import net.resnet_euc as resnet
        net = resnet.ResNet34(num_c=args.class_num)
    elif args.net=='resnet_dropout34':
        import net.resnet_dropout as resnet
        net = resnet.ResNet34(num_c=args.class_num)
    elif args.net=='resnet_dropout50':
        import net.resnet_dropout as resnet
        net = resnet.ResNet50(num_c=args.class_num)
    net.load_state_dict(torch.load(model_path)['state_dict'])
    if torch.cuda.device_count()>1:
        net=torch.nn.DataParallel(net)
    return net.cuda()
def print_info(args):
    type_list=['eval','ood']
    for t in type_list:
        if args.save.find('ensemble')>0:
            info=torch.load('{}/{}_info_{}.pt'.format(args.save,t,args.calibration))
            with open('{}/{}_info_{}.txt'.format(args.save,t,args.calibration),'wt') as f:
                for item in info:
                    print('{}: {}'.format(item,info[item]),file=f)
        else:
            info=torch.load('{}/{}_info.pt'.format(args.save,t))
            with open('{}/{}_info.txt'.format(args.save,t),'wt') as f:
                for item in info:
                    print('{}: {:.4f}'.format(item,info[item]),file=f)
 
def write_to_excel(args):
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    print(args)
    out_datasets=['out','val','Imagenet','Imagenet_resize','LSUN','LSUN_resize','iSUN']

    first_column_keys=['']
    for item in out_datasets:
        first_column_keys.append('{}_{}_tnr95'.format(args.in_dataset,item))
        first_column_keys.append('{}_{}_auroc'.format(args.in_dataset,item))
    table_list=['Pred','Pred_Tsc']

    if 'eval' in args.inferance:
        max_calibration=round(torch.load('{}/calibration.pt'.format(args.save))['eval'],2)
        print(max_calibration)
        # calibration_list=[max_calibration]
        # calibration_list.append(0.0)
        # calibration_list.append(1.0)
        calibration_list=[round(i*0.05,2) for i in range(21)]
        wb = Workbook()
        ws = wb['Sheet']
        wb.remove(ws)
        
        for item in table_list:
            wb.create_sheet(item)
            ws = wb[item]
            for i in range(len(first_column_keys)):
                ws.cell(row=i+1,column=1).value=first_column_keys[i]
            ws.column_dimensions[get_column_letter(1)].width=25
            
            for j in range(len(calibration_list)):
                info=torch.load('{}/eval_info_{}.pt'.format(args.save,calibration_list[j]))
                ws.cell(row=1,column=j+2).value=calibration_list[j]
                for i in range(len(first_column_keys)-1):
                    ws.cell(row=i+2,column=j+2).value=round(info['{}_cali_{}'.format(first_column_keys[i+1],item)].item()*100,1)
        save_file='../models/{}/eval_{}_{}.xlsx'.format(args.save,args.in_dataset,args.ood_batch_size)
        wb.save(save_file)


    if 'ood' in args.inferance:
        max_calibration=round(torch.load('{}/calibration.pt'.format(args.save))['ood'],2)
        # calibration_list=[max_calibration]
        # calibration_list.append(0.0)
        # calibration_list.append(1.0)
        calibration_list=[round(i*0.05,2) for i in range(21)]
        wb = Workbook()
        ws = wb['Sheet']
        wb.remove(ws)
        
        for item in table_list:
            wb.create_sheet(item)
            ws = wb[item]
            for i in range(len(first_column_keys)):
                ws.cell(row=i+1,column=1).value=first_column_keys[i]
            ws.column_dimensions[get_column_letter(1)].width=25
            
            for j in range(len(calibration_list)):
                info=torch.load('{}/ood_info_{}.pt'.format(args.save,calibration_list[j]))
                ws.cell(row=1,column=j+2).value=calibration_list[j]
                for i in range(len(first_column_keys)-1):
                    ws.cell(row=i+2,column=j+2).value=round(info['{}_cali_{}'.format(first_column_keys[i+1],item)].item()*100,1)
        save_file='../models/{}/ood_{}_{}.xlsx'.format(args.save,args.in_dataset,args.ood_batch_size)
        wb.save(save_file)

def plt_fig(args):
    out_dataset_list=['Imagenet','Imagenet_resize', 'LSUN', 'LSUN_resize', 'iSUN', 'SVHN']
    noise_dataset_list=['Gaussian', 'Uniform']
    st_list=['Pred','Pred_Tsc','cali_Pred','cali_Pred_Tsc','bias_Pred','bias_Pred_Tsc']
    od=out_dataset_list[1]
    nd=noise_dataset_list[0]
    st=st_list[5]
    
    in_score=torch.load('{}/score/{}_out_{}.pt'.format(args.save,args.in_dataset,st))
    out_score=torch.load('{}/score/{}_{}.pt'.format(args.save,od,st))
    noise_score=torch.load('{}/score/{}_{}.pt'.format(args.save,nd,st))

    # in_score=-torch.sum(in_score*torch.log(in_score),dim=1)
    # out_score=-torch.sum(out_score*torch.log(out_score),dim=1)
    # noise_score=-torch.sum(noise_score*torch.log(noise_score),dim=1)

    # min_value=torch.min(torch.min(in_score),torch.min(torch.min(out_score),torch.min(noise_score))).item()
    # in_score,out_score,noise_score=(in_score-min_value)*1e6,(out_score-min_value)*1e6,(noise_score-min_value)*1e6


    plt_data=[in_score,out_score,noise_score]
    max_value=torch.max(torch.max(in_score),torch.max(torch.max(out_score),torch.max(noise_score))).item()
    min_value=torch.min(torch.min(in_score),torch.min(torch.min(out_score),torch.min(noise_score))).item()
    print(max_value,min_value)
    import matplotlib.pyplot as plt

    
    fig=plt.figure(figsize=(8,5))


    color_brief=['r','b','g','y','c','m','w']
    

    for i in range(len(plt_data)):
        plt.hist(plt_data[i].detach().cpu().numpy(), bins=1000,range=(min_value,max_value), density=False, histtype='bar',\
            color=color_brief[i],alpha=1-0.3*i,log=True)
    plt.show()
    return

def create_TinyImagenet_val_img_folder():
    '''
    This method is responsible for separating validation images into separate sub folders
    '''
    dataset_dir = '../data/TinyImagenet'
    val_dir = '../data/TinyImagenet/val'
    img_dir = '{}/images'.format(val_dir)

    fp = open('../data/TinyImagenet/val/val_annotations.txt', 'r')
    data = fp.readlines()
    val_img_dict = {}
    for line in data:
        words = line.split('\t')
        val_img_dict[words[0]] = words[1]
    fp.close()

    # Create folder if not present and move images into proper folders
    for img, folder in val_img_dict.items():
        newpath = ('{}/{}'.format(img_dir, folder))
        if not os.path.exists(newpath):
            os.makedirs(newpath)
        if os.path.exists('{}/{}'.format(img_dir, img)):
            os.rename('{}/{}'.format(img_dir, img), '{}/{}'.format(newpath, img))
