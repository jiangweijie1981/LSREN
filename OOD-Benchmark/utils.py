import numpy as np
import random
import torch


def set_seed(seed):
    np.random.seed(seed)
    random.seed(seed)
    torch.manual_seed(seed) # cpu
    torch.cuda.manual_seed_all(seed)  # gpu
    torch.backends.cudnn.benchmark=False
    torch.backends.cudnn.deterministic = True

def get_smooth_list(args):
    smooth_list=[]
    if args.ensemble_num==1:
        smooth_list.append(0.0)
    elif args.ensemble_num==2:
        smooth_list.append(args.smooth_range)
    else:
        for i in range(args.ensemble_num-2):
            smooth_list.append(round(args.smooth_range*(i+1)/(args.ensemble_num-1),6))
    return smooth_list

def get_ensemble_list(args):
    model_file_list=[]
    smooth_list=get_smooth_list(args)
    if args.train_type=='LSR':
        for i in range(len(smooth_list)):
            args.smooth=smooth_list[i]
            file_path = '../models/{}_seed{}_{}_{}{}/{}_esbl{}_range{}/epochs{}_s{:.6f}'.format(
                args.net,args.seed,args.in_dataset,args.noise_type,args.noise_ratio,
                args.train_type,args.ensemble_num,args.smooth_range,args.epochs,args.smooth)
            model_file_list.append(file_path)
        if args.ensemble_num>1:
            args.smooth=0.0
            file_path = '../models/{}_seed{}_{}_{}{}/{}_esbl1_range{}/epochs{}_s{:.6f}'.format(
                args.net,args.seed,args.in_dataset,args.noise_type,args.noise_ratio,
                args.train_type,args.smooth_range,args.epochs,args.smooth)
            model_file_list.append(file_path)
        if args.ensemble_num>2:
            args.smooth=args.smooth_range
            file_path = '../models/{}_seed{}_{}_{}{}/{}_esbl2_range{}/epochs{}_s{:.6f}'.format(
                args.net,args.seed,args.in_dataset,args.noise_type,args.noise_ratio,
                args.train_type,args.smooth_range,args.epochs,args.smooth)
            model_file_list.append(file_path)
    if args.train_type=='ILSR':
        for i in range(len(smooth_list)):
            args.smooth=smooth_list[i]
            file_path = '../models/{}_seed{}_{}_{}{}/{}_esbl{}_range{}_penalty{}/epochs{}_s{:.6f}'.format(
                args.net,args.seed,args.in_dataset,args.noise_type,args.noise_ratio,
                args.train_type,args.ensemble_num,args.smooth_range,args.penalty,args.epochs,args.smooth)
            model_file_list.append(file_path)
        if args.ensemble_num>1:
            args.smooth=0.0
            file_path = '../models/{}_seed{}_{}_{}{}/{}_esbl1_range{}_penalty{}/epochs{}_s{:.6f}'.format(
                args.net,args.seed,args.in_dataset,args.noise_type,args.noise_ratio,
                args.train_type,args.smooth_range,args.penalty,args.epochs,args.smooth)
            model_file_list.append(file_path)
        if args.ensemble_num>2:
            args.smooth=args.smooth_range
            file_path = '../models/{}_seed{}_{}_{}{}/{}_esbl2_range{}_penalty{}/epochs{}_s{:.6f}'.format(
                args.net,args.seed,args.in_dataset,args.noise_type,args.noise_ratio,
                args.train_type,args.smooth_range,args.penalty,args.epochs,args.smooth)
            model_file_list.append(file_path)
    return model_file_list
def get_net(args):
    if args.net=='dense':
        import net.densenet as densenet
        net = densenet.DenseNet(growthRate=12, depth=100, reduction=0.5,bottleneck=True, nClasses=args.class_num)
    elif args.net=='wrn':
        import net.wrn as wideresnet
        net = wideresnet.WideResNet(depth=28,num_classes=args.class_num,widen_factor=10,dropRate=0)
    elif args.net=='resnet':
        import net.resnet as resnet
        net = resnet.ResNet34(num_c=args.class_num)
    if torch.cuda.device_count()>1:
        net=torch.nn.DataParallel(net)
    return net.cuda()
def load_net(args,model_path):
    if args.net=='dense':
        import net.densenet as densenet
        net = densenet.DenseNet(growthRate=12, depth=100, reduction=0.5,bottleneck=True, nClasses=args.class_num)
    elif args.net=='wrn':
        import net.wrn as wideresnet
        net = wideresnet.WideResNet(depth=28,num_classes=args.class_num,widen_factor=10,dropRate=0)
    elif args.net=='resnet':
        import net.resnet as resnet
        net = resnet.ResNet34(num_c=args.class_num)
    net.load_state_dict(torch.load(model_path)['state_dict'])
    if torch.cuda.device_count()>1:
        net=torch.nn.DataParallel(net)
    return net.cuda()
def print_info(args):
    if args.save.find('ensemble')>0:
        info=torch.load('{}/info_{}.pt'.format(args.save,args.calibration))
    else:
        info=torch.load('{}/info.pt'.format(args.save))
    metric_type_list=['tnr95','auroc']
    out_dataset_list=['Imagenet','Imagenet_resize', 'LSUN', 'LSUN_resize', 'iSUN', 'SVHN','Gaussian', 'Uniform']
    if args.in_dataset=='cifar10':
        out_dataset_list.append('cifar100')
    elif args.in_dataset=='cifar100':
        out_dataset_list.append('cifar10')
    if args.save.find('ensemble')>0:
        with open('{}/info_{}.txt'.format(args.save,args.calibration),'wt') as f:
            print('model:{}\n'.format(args.save),file=f)
            print('magnitude:{}\n'.format(info['magnitude']),file=f)
            print('\t\tPred\tPred_cali\tPred_Tsc\tPred_Tsc_cali\n',file=f)
            for mt in metric_type_list:
                for od in out_dataset_list:
                    print("{}_{}\t\t{:.2f}%\t{:.2f}%\t{:.2f}%\t{:.2f}%".format(
                            od,mt,
                            info['{}_{}_Pred'.format(od,mt)]*100,
                            info['{}_{}_cali_Pred'.format(od,mt)]*100,
                            info['{}_{}_Pred_Tsc'.format(od,mt)]*100,
                            info['{}_{}_cali_Pred_Tsc'.format(od,mt)]*100,
                            ),file=f)
    else:
        with open('{}/info.txt'.format(args.save),'wt') as f:
            print('model:{}\n'.format(args.save),file=f)
            print('magnitude:{}\n'.format(info['magnitude']),file=f)
            print('\t\tPred\tPred_Tsc',file=f)
            for mt in metric_type_list:
                for od in out_dataset_list:
                    print("{}_{}\t\t{:.2f}%\t{:.2f}%".format(
                            od,mt,
                            info['{}_{}_Pred'.format(od,mt)]*100,
                            info['{}_{}_Pred_Tsc'.format(od,mt)]*100,
                            ),file=f)


def write_to_excel(args):
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    save_file='../models/{}/{}-{}-{}-{}-{}.xlsx'.format(args.save,args.net,args.in_dataset,args.seed,args.ood_batch_size,args.infer)
    if args.in_dataset=='cifar10':
        first_column_keys=['','Imagenet_tnr95','Imagenet_resize_tnr95', 'LSUN_tnr95', 'LSUN_resize_tnr95',\
             'iSUN_tnr95','SVHN_tnr95', 'Gaussian_tnr95', 'Uniform_tnr95','cifar100_tnr95','',\
            'Imagenet_auroc','Imagenet_resize_auroc', 'LSUN_auroc', 'LSUN_resize_auroc',\
                 'iSUN_auroc','SVHN_auroc', 'Gaussian_auroc', 'Uniform_auroc','cifar100_auroc']
    elif args.in_dataset=='cifar100':
        first_column_keys=['','Imagenet_tnr95','Imagenet_resize_tnr95', 'LSUN_tnr95', 'LSUN_resize_tnr95',\
             'iSUN_tnr95','SVHN_tnr95', 'Gaussian_tnr95', 'Uniform_tnr95','cifar10_tnr95','',\
            'Imagenet_auroc','Imagenet_resize_auroc', 'LSUN_auroc', 'LSUN_resize_auroc',\
                 'iSUN_auroc','SVHN_auroc', 'Gaussian_auroc', 'Uniform_auroc','cifar10_auroc']
    table_list=['Pred','Pred_Tsc']
    calibration_list=[]
    count=int(1/args.calibration_interval)
    for i in range(count+1):
        calibration_list.append(round(i*args.calibration_interval,2))
    # calibration_list=[0.25,0.3]

    wb = Workbook()
    ws = wb['Sheet']
    wb.remove(ws)
    
    for item in table_list:
        wb.create_sheet(item)
        ws = wb[item]
        for i in range(len(first_column_keys)):
            ws.cell(row=i+1,column=1).value=first_column_keys[i]
        ws.column_dimensions[get_column_letter(1)].width=25
        
        for j in range(len(calibration_list)):
            info=torch.load('{}/info_{}.pt'.format(args.save,calibration_list[j]))
            ws.cell(row=1,column=j+2).value=calibration_list[j]
            for i in range(len(first_column_keys)-1):
                if i==9:
                    continue
                ws.cell(row=i+2,column=j+2).value=round(info['{}_cali_{}'.format(first_column_keys[i+1],item)].item()*100,1)
    
    wb.save(save_file)
 

def plt_fig(args):
    out_dataset_list=['Imagenet','Imagenet_resize', 'LSUN', 'LSUN_resize', 'iSUN', 'SVHN']
    noise_dataset_list=['Gaussian', 'Uniform']
    st_list=['Pred','Pred_Tsc','cali_Pred','cali_Pred_Tsc','bias_Pred','bias_Pred_Tsc']
    od=out_dataset_list[1]
    nd=noise_dataset_list[0]
    st=st_list[5]
    
    in_score=torch.load('{}/score/{}_out_{}.pt'.format(args.save,args.in_dataset,st))
    out_score=torch.load('{}/score/{}_{}.pt'.format(args.save,od,st))
    noise_score=torch.load('{}/score/{}_{}.pt'.format(args.save,nd,st))

    # in_score=-torch.sum(in_score*torch.log(in_score),dim=1)
    # out_score=-torch.sum(out_score*torch.log(out_score),dim=1)
    # noise_score=-torch.sum(noise_score*torch.log(noise_score),dim=1)

    # min_value=torch.min(torch.min(in_score),torch.min(torch.min(out_score),torch.min(noise_score))).item()
    # in_score,out_score,noise_score=(in_score-min_value)*1e6,(out_score-min_value)*1e6,(noise_score-min_value)*1e6


    plt_data=[in_score,out_score,noise_score]
    max_value=torch.max(torch.max(in_score),torch.max(torch.max(out_score),torch.max(noise_score))).item()
    min_value=torch.min(torch.min(in_score),torch.min(torch.min(out_score),torch.min(noise_score))).item()
    print(max_value,min_value)
    import matplotlib.pyplot as plt

    
    fig=plt.figure(figsize=(8,5))


    color_brief=['r','b','g','y','c','m','w']
    

    for i in range(len(plt_data)):
        plt.hist(plt_data[i].detach().cpu().numpy(), bins=1000,range=(min_value,max_value), density=False, histtype='bar',\
            color=color_brief[i],alpha=1-0.3*i,log=True)
    plt.show()
    return
