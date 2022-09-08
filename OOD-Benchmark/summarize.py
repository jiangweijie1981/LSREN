from torch import seed
from configs import *
from utils import *
import os
from cal_metric import *
import shutil
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


if __name__=='__main__':
    args=Config().data
    ood_aur1,ood_tnr1=[],[]
    ood_aur2,ood_tnr2=[],[]
    table_list=['Pred','Pred_Tsc']
    for s in [0,1,2]:
        args.seed=s
        set_seed(args.seed)
        ensemble_path = '../models/{}_seed{}_{}_{}{}/{}_esbl{}_range{}/ensemble_epochs{}'.format(
            args.net,args.seed,args.in_dataset,args.noise_type,args.noise_ratio,
            args.train_type,args.ensemble_num,args.smooth_range,args.epochs)

        wb = load_workbook('{}/{}-{}-{}-{}.xlsx'.format(ensemble_path,args.net,args.in_dataset,args.seed,args.ood_batch_size,args.infer))
        ws = wb['Pred_Tsc']

        max_idx,max_value=1,ws[10][1].value+ws[20][1].value
        for i in range(21):
            if args.in_dataset=='cifar4':
                cur_value=ws[6][i+1].value+ws[7][i+1].value
            else:
                cur_value=ws[4][i+1].value+ws[5][i+1].value
            if cur_value>max_value:
                max_idx=i+1
                max_value=cur_value
        
        if args.in_dataset=='cifar4':
            ood_aur.append(ws[4][max_idx].value)
            ood_aur1.append(ws[5][max_idx].value)
        else:
            ood_aur.append(ws[3][max_idx].value)

    
    with open('summarize_{}_{}.txt'.format(args.in_dataset,args.infer),'wt') as f:
        print('ood_aur:{:.1f}\n'.format(np.mean(ood_aur)),file=f)
        print('ood_aurs:',file=f)
        print(ood_aur,file=f)
        if args.in_dataset=='cifar4':
            print('ood_aur1:{:.1f}\n'.format(np.mean(ood_aur1)),file=f)
            print('ood_aurs1:',file=f)
            print(ood_aur1,file=f)





    with open('summarize_en{}_{}_{}.txt'.format(args.ensemble_num,args.ood_batch_size,args.inferance),'wt') as f:
        print('Pred\n',file=f)
        print('ood_aur:{:.1f}\n'.format(np.mean(ood_aur1)),file=f)
        print('ood_tnr:{:.1f}\n'.format(np.mean(ood_tnr1)),file=f)

        print('Pred_Tsc\n',file=f)
        print('ood_aur:{:.1f}\n'.format(np.mean(ood_aur2)),file=f)
        print('ood_tnr:{:.1f}\n'.format(np.mean(ood_tnr2)),file=f)



